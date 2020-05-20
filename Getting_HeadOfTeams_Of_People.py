#coding: utf-8
"""
Created on Mon Nov  4 10:49:55 2019

@author: Antonio Raffaele Iannaccone
"""

import tkinter
from tkinter.filedialog import askopenfilename
from tkinter import filedialog
from selenium import webdriver
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import pyautogui
import datetime
import os

global cwd
cwd = os.getcwd()
print(cwd)

#Define function for choosing the Excel file
def chooseExcel():
    global filePath
    filePath = askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")]) # show an "Open" dialog box and return the path to the selected file
    filePath = filePath.replace("/", "\\")

#Define function for choosing the output location
def chooseOutputLocation():
    global outputPath
    #Ask for a directory where to put the outputted Excel file
    outputPath = filedialog.askdirectory()
    outputPath = outputPath.replace("/", "\\")

#Define main function
def mainFunction():
    global filePath
    global outputPath
    global cwd
    
    #Check if an Excel file was selected or not
    if 'filePath' not in globals():
        pyautogui.alert(title='Information', text='Before using this program, please, choose your excel sheet (with names in column B) and the output location.')
        return
    
    if 'outputPath' not in globals():
        pyautogui.alert(title='Information', text='Before using this program, please, choose your excel sheet (with names in column B) and the output location.')
        return
    
    try:
        #WORKING WITH EXCEL
        wb = load_workbook(filename=str(filePath))
        ws = wb['Sheet1']
        row_count = ws.max_row
        column_count = ws.max_column
        print('Row count is: ' + str(row_count))
        print('Column count is: ' + str(column_count))
        
    except:
        pyautogui.alert(title='Error', text='Something went wrong during the initialization of the existing Excel file.')
    
    #Initiate Selenium webdriver
    driver = webdriver.Chrome(cwd + '\\chromedriver.exe') 
    driver.get("https://UrlCanNotBeDisclosed.com")
    
    try:
        wb_output = Workbook(write_only=True)
        ws_output = wb_output.create_sheet(title='Names with teamleads')
        
        #Define the rowIdentificator variable
        rowIdentificator = 1
    except:
        print('Something went wrong during the initalization of the output Excel sheet.')
    
    
    for cell in ws['B']:
        print(cell.value + "#")
        
        search = driver.find_element_by_id("search")
        search.send_keys(str(cell.value))
        
        search_button = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/nav/div[2]/form/div/span/button')
        search_button.click()
        
        time.sleep(0.7)
        
        try:
            line_manager_button = driver.find_element_by_xpath('//*[@id="pt-box-2"]/div[3]/div[2]/a')
            line_manager_button.click()
        except:
            
            try:
                no_employees_found = driver.find_element_by_xpath('//*[@id="content"]/div[2]/div[3]/div[1]/div[2]/div')
                manager = "Multiple employees found."
            
                print('Multiple employees found.')
                empty_search = driver.find_element_by_id("search")
                empty_search.clear()
            
                ws_output.append([str(cell.value), str(manager)])
                continue
            
            except:
                manager = "No employees found."
            
                print('No employees found.')
                empty_search = driver.find_element_by_id("search")
                empty_search.clear()
            
                ws_output.append([str(cell.value), str(manager)])
                continue
        
        time.sleep(0.7)
        line_manager_email = driver.find_element_by_xpath('//*[@id="pt-box-1"]/div[4]/div[2]/a[1]')
        print(line_manager_email.text + "#")
        
        manager = line_manager_email.text
                      
        empty_search = driver.find_element_by_id("search")
        empty_search.clear()
        
        ws_output.append([str(cell.value), str(manager)])
        
        rowIdentificator +=1
        manager = ""
        
        print("")
    
    
    ###Add the time stamp to the name of file and save the outputted Excel file to the selected location
    try:
        actualTime = time.time()
        convertedActualTime = datetime.datetime.fromtimestamp(actualTime).strftime('%d_%m_%Y-%H_%M_%S')
        savedPathOfOutputtedFile = (str(outputPath) + r'\Names_Of_HeadOfTeams_' + str(convertedActualTime) + '.xlsx')
        wb_output.save(str(savedPathOfOutputtedFile))
        pyautogui.alert(title='Information', text="The outputted excel sheet has been saved in the selected location. It's name starts with Names_Of_HeadOfTeams_xxxxx.xlsx.")
    except:
        pyautogui.alert(title='Error', text='Something went wrong during the saving of the excel sheet. Please, try again.')


#GUI        
top = tkinter.Tk()

top.title('Get Names of TLs')
top.geometry('340x260')

button = tkinter.Button(text = 'Start!', command = mainFunction)
button.grid(row = 1, column=1, columnspan=6)

buttonChoose = tkinter.Button(text = 'Choose Excel with Names', command = chooseExcel)
buttonChoose.grid(row = 2, column=1, columnspan=6)

buttonChooseOutputLocation = tkinter.Button(text = 'Choose Output Location', command = chooseOutputLocation)
buttonChooseOutputLocation.grid(row = 3, column=1, columnspan=6)

label1 = tkinter.Label(text = 'Important note:\nIn the selected excel sheet, the names\nof the people have to be in the column B.')
label1.grid(row = 4, column=1, columnspan=6)

labelX = tkinter.Label(text = '')
labelX.grid(row = 5, column=1, columnspan=6)

label2 = tkinter.Label(text = 'Created by\nAntonio Raffaele Iannaccone')
label2.grid(row = 6, column=1, columnspan=6)

top.mainloop()